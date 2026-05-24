import openpyxl
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from difflib import get_close_matches
from typing import Optional
from .models import ObjectOfExpenditure, ObjectCode, ItemCode, Item

VALID_UNITS = {choice[0] for choice in Item.Unit.choices}

@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list = field(default_factory=list)


def normalize_unit(raw_unit: str) -> Optional[str]:
    """
    Normalizes a raw Excel unit string to a valid Unit choice.
    Handles whitespace, plurals, and common typos via fuzzy matching.
    Returns None if no confident match is found.
    """
    if not raw_unit:
        return None

    cleaned = str(raw_unit).strip().lower()

    # 1. Exact match
    if cleaned in VALID_UNITS:
        return cleaned

    # 2. Depluralize — handles "pieces", "reams", "gallons" etc.
    depluralized = cleaned.rstrip("s")
    if depluralized in VALID_UNITS:
        return depluralized

    # 3. Fuzzy match — handles typos like "botlles", "peice"
    matches = get_close_matches(cleaned, VALID_UNITS, n=1, cutoff=0.75)
    if matches:
        return matches[0]

    return None


def parse_unit_cost(raw_cost) -> Optional[Decimal]:
    try:
        return Decimal(str(raw_cost)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return None


def import_items_from_excel(file) -> ImportResult:
    result = ImportResult()

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        result.errors.append(f"Failed to read file: {e}")
        return result

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    required_columns = {
        "Item", "Specification", "Unit", "Unit Cost",
        "General Descriptions", "Code", "Object Code", "Object Expenditure"
    }

    if not required_columns.issubset(set(headers)):
        missing = required_columns - set(headers)
        result.errors.append(f"Missing required columns: {missing}")
        return result

    column = {name: idx for idx, name in enumerate(headers)}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if not any(row):
            continue

        def get(column_name):
            val = row[column[column_name]]
            return str(val).strip() if val is not None else ""

        expenditure_name = get("Object Expenditure")
        object_code_value = get("Object Code")
        item_code_value = get("Code")
        general_description = get("General Descriptions")
        item_name = get("Item")
        specification = get("Specification") or ""
        raw_unit = get("Unit")
        raw_cost = row[column["Unit Cost"]]

        # Validate required fields
        if not all([expenditure_name, object_code_value, item_code_value,
                    general_description, item_name, raw_unit]):
            result.errors.append(
                f"Row {row_num}: Missing required field(s) — skipped."
            )
            result.skipped += 1
            continue

        unit = normalize_unit(raw_unit)
        if not unit:
            result.errors.append(
                f"Row {row_num}: Unrecognized unit '{raw_unit}' for item '{item_name}' — skipped."
            )
            result.skipped += 1
            continue

        unit_cost = parse_unit_cost(raw_cost)
        if unit_cost is None or unit_cost <= 0:
            result.errors.append(
                f"Row {row_num}: Invalid unit cost '{raw_cost}' for item '{item_name}' — skipped."
            )
            result.skipped += 1
            continue

        # get_or_create hierarchy

        expenditure, _ = ObjectOfExpenditure.objects.get_or_create(
            name=expenditure_name
        )

        object_code, _ = ObjectCode.objects.get_or_create(
            code=object_code_value,
            defaults={"expenditure": expenditure}
        )

        item_code, _ = ItemCode.objects.get_or_create(
            code=item_code_value,
            object_code=object_code,
            defaults={"general_description": general_description}
        )

        # Update general_description if it changed
        if item_code.general_description != general_description:
            item_code.general_description = general_description
            item_code.save()

        existing = Item.objects.filter(
            name=item_name,
            item_code=item_code,
            specification=specification      
        ).first()

        if existing:
            existing.unit = unit
            existing.unit_cost = unit_cost
            existing.save()
            result.updated += 1
        else:
            Item.objects.create(
                name=item_name,
                item_code=item_code,
                specification=specification,
                unit=unit,
                unit_cost=unit_cost,
            )
            result.created += 1

    wb.close()
    return result