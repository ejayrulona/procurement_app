"""
Project Procurement Management Plan (PPMP) - Excel Export Service

Calibrated to the official NGPA PPMP Template (RA 12009).

Column mapping (data rows, cols A–N):
  A  = General Description and Objective of the Project to be Procured
  B  = Type of the Project (Classification: Goods / Infrastructure / Consulting Services)
  C  = Quantity
  D  = Unit
  E  = Description / Specification
  F  = Recommended Mode of Procurement
  G  = Pre-Procurement Conference, if applicable
  H  = Start of Procurement Activity (MM/YYYY)
  I  = End of Procurement Activity (MM/YYYY)
  J  = Expected Delivery / Implementation Period (MM/YYYY)
  K  = Source of Funds
  L  = Estimated Budget / Authorized Budgetary Allocation (PhP)
  M  = Attached Supporting Documents
  N  = Remarks
"""

import io
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

# CONSTANTS

DATA_START_ROW = 19   # First data row in the template (rows 1–18 are headers/instructions)
TOTAL_BUDGET_LABEL = "TOTAL BUDGET:"

CURRENCY_FMT = '#,##0.00'
DATE_FMT = 'MM/YYYY'


# PUBLIC API

def generate_ppmp_excel(
    records: list[dict],
    template_path: str,
    fiscal_year: int | None = None,
    end_user_unit: str | None = None,
    ppmp_number: str | None = None,
    submission_type: str | None = None,
) -> bytes:
    """
    Populate the official PPMP Excel template with procurement records.

    Steps:
      1. Load the template (preserves all formatting, merged cells, headers)
      2. Fill in header metadata (fiscal year, end-user unit, PPMP number)
      3. Clear existing data rows (values only — formatting stays intact)
      4. Write new records starting at DATA_START_ROW
      5. Add TOTAL BUDGET formula below the last data row
      6. Serialize to bytes and return

    Parameters
    records : list[dict]
        Each dict should contain:

        {
            "general_description":     str,   # Full description/objective of the project
            "classification":          str,   # "Goods" | "Infrastructure" | "Consulting Services"
            "quantity":                int | float,
            "unit":                    str,   # e.g. "lot", "pcs", "reams"
            "specification":           str,   # Technical description / specification
            "mode_of_procurement":     str,   # Display label of mode
            "start_of_procurement":    date | datetime | str | None,  # MM/YYYY
            "end_of_procurement":      date | datetime | str | None,  # MM/YYYY
            "delivery_date":           date | datetime | str | None,  # MM/YYYY
            "source_of_funds":         str,
            "estimated_budget":        float | int | None,
            "remarks":                 str | None,
        }

        Columns G (Pre-Procurement Conference) and M (Attached Supporting Documents)
        are intentionally left blank for the admin to fill in manually.

    template_path : str
        Absolute path to the official .xlsx template file.

    fiscal_year : int | None
        Fiscal year to write into the header row.

    end_user_unit : str | None
        End-user / implementing unit name for the header.

    ppmp_number : str | None
        PPMP number to write into the title row.

    submission_type : str | None
        "Indicative" or "Final" — written into the PPMP type header row.

    Returns
    -------
    bytes
        Raw .xlsx bytes ready to stream as an HTTP file-download response.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    _write_header_metadata(ws, fiscal_year, end_user_unit, ppmp_number, submission_type)
    _clear_data_rows(ws, DATA_START_ROW, ws.max_row)

    for i, record in enumerate(records):
        _write_record_row(ws, DATA_START_ROW + i, record)

    # Add total budget formula one row below the last data row
    last_data_row = DATA_START_ROW + len(records) - 1
    total_row = last_data_row + 1
    _write_total_row(ws, total_row, DATA_START_ROW, last_data_row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# HEADER METADATA

def _write_header_metadata(ws, fiscal_year, end_user_unit, ppmp_number, submission_type):
    """Write fiscal year, end-user unit, PPMP number, and submission type into header cells."""

    # Scan for header label cells and write adjacent values
    for row in ws.iter_rows(max_row=DATA_START_ROW - 1):
        for cell in row:
            val = str(cell.value or "").strip()

            if "Fiscal Year" in val and fiscal_year is not None:
                # Write fiscal year into the cell itself or adjacent
                cell.value = f"Fiscal Year : {fiscal_year}"

            elif "End-User or Implementing Unit" in val and end_user_unit:
                cell.value = f"End-User or Implementing Unit: {end_user_unit}"

            elif "PROJECT PROCUREMENT MANAGEMENT PLAN (PPMP) NO." in val:
                suffix = f", for CY {fiscal_year}" if fiscal_year else ""
                num = ppmp_number or "___"
                cell.value = f"PROJECT PROCUREMENT MANAGEMENT PLAN (PPMP) NO. {num}{suffix}"

            elif "PPMP Type" in val and submission_type:
                cell.value = f"PPMP Type: {submission_type}"


# CLEAR — wipe values only, preserve all cell formatting

def _clear_data_rows(ws, start_row: int, end_row: int):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


# WRITE — one record → one row

def _write_record_row(ws, row: int, rec: dict):
    # A: General Description and Objective
    _set(ws, f"A{row}", rec.get("general_description"), wrap=True)

    # B: Classification (Type of Project)
    _set(ws, f"B{row}", rec.get("classification"))

    # C: Quantity
    qty = rec.get("quantity")
    if qty is not None:
        try:
            ws[f"C{row}"].value = float(qty)
        except (ValueError, TypeError):
            ws[f"C{row}"].value = qty

    # D: Unit of Measure
    _set(ws, f"D{row}", rec.get("unit"))

    # E: Description / Specification
    _set(ws, f"E{row}", rec.get("specification"), wrap=True)

    # F: Recommended Mode of Procurement
    _set(ws, f"F{row}", rec.get("mode_of_procurement"), wrap=True)

    # G: Pre-Procurement Conference — intentionally left blank (admin fills in Excel)

    # H: Start of Procurement Activity (MM/YYYY)
    _set_date(ws, f"H{row}", rec.get("start_of_procurement"))

    # I: End of Procurement Activity (MM/YYYY)
    _set_date(ws, f"I{row}", rec.get("end_of_procurement"))

    # J: Expected Delivery / Implementation Period (MM/YYYY)
    _set_date(ws, f"J{row}", rec.get("delivery_date"))

    # K: Source of Funds
    _set(ws, f"K{row}", rec.get("source_of_funds"), wrap=True)

    # L: Estimated Budget / Authorized Budgetary Allocation
    _set_currency(ws, f"L{row}", rec.get("estimated_budget"))

    # M: Attached Supporting Documents — intentionally left blank (admin fills in Excel)

    # N: Remarks
    _set(ws, f"N{row}", rec.get("remarks"), wrap=True)


def _write_total_row(ws, total_row: int, data_start: int, data_end: int):
    ws[f"K{total_row}"] = TOTAL_BUDGET_LABEL
    ws[f"L{total_row}"] = f"=SUM(L{data_start}:L{data_end})"
    ws[f"L{total_row}"].number_format = CURRENCY_FMT


# CELL HELPERS

def _set(ws, addr: str, value, wrap: bool = False):
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        return
    cell.value = None if value in (None, "", "—", "–", "--") else str(value)
    if wrap:
        existing = cell.alignment
        cell.alignment = Alignment(
            wrap_text=True,
            vertical=getattr(existing, "vertical", None) or "top",
            horizontal=getattr(existing, "horizontal", None) or "left",
        )


def _set_date(ws, addr: str, value):
    """Write a date/datetime. Applies MM/YYYY number format to match template style."""
    cell = ws[addr]
    if value in (None, "", "—", "–", "--"):
        cell.value = None
        return
    if isinstance(value, datetime):
        cell.value = value
    elif isinstance(value, date):
        cell.value = datetime(value.year, value.month, value.day)
    elif isinstance(value, str):
        try:
            cell.value = datetime.fromisoformat(value)
        except ValueError:
            cell.value = value
    else:
        cell.value = value
    cell.number_format = DATE_FMT


def _set_currency(ws, addr: str, value):
    cell = ws[addr]
    if value in (None, ""):
        cell.value = None
        return
    try:
        cell.value = float(value)
    except (ValueError, TypeError):
        cell.value = None
    cell.number_format = CURRENCY_FMT